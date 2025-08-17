import napari
import numpy as np
import nibabel as nib
import pandas as pd
import os
import pickle
import yaml
from yaml.loader import SafeLoader
import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import hashlib

def load_nifti_image(path):
    """Loads a NIfTI (.nii) image and returns it as a NumPy array."""
    img = nib.load(path)
    return img.get_fdata()

def load_best_models_from_csv(csv_path="best_models.csv"):
    """Load the best models from the CSV file."""
    try:
        df = pd.read_csv(csv_path)
        print(f"Loaded {len(df)} best models from {csv_path}")
        return df
    except Exception as e:
        print(f"Error loading CSV file {csv_path}: {e}")
        return None

def find_pickle_file_for_model(row):
    """Find the corresponding pickle file for a model row."""
    # Check if pickle path is directly provided in the CSV
    if 'pickle' in row and pd.notna(row['pickle']) and os.path.exists(row['pickle']):
        print(f"Found pickle file from CSV: {row['pickle']}")
        return row['pickle']
    
    # Fallback to original search logic if pickle column is not available or file doesn't exist
    # Extract key identifiers from the paths
    config_path = row['config_path']
    scan_path = row['mct_path']
    
    # Get the config name and scan name
    config_name = os.path.splitext(os.path.basename(config_path))[0]
    scan_name = os.path.splitext(os.path.basename(scan_path).replace('-image.nii', ''))[0]
    
    # If y_hat_path is available, use it for more precise matching
    if 'y_hat_path' in row and pd.notna(row['y_hat_path']):
        y_hat_path = row['y_hat_path']
        # Extract the base name and convert from .csv to .pickle
        y_hat_basename = os.path.basename(y_hat_path)
        if y_hat_basename.endswith('.csv'):
            # Remove .resampled_space_peaks.csv and add _results.pickle
            pickle_basename = y_hat_basename.replace('.resampled_space_peaks.csv', '_results.pickle')
        else:
            pickle_basename = y_hat_basename.replace('.csv', '_results.pickle')
        
        # Look in analysis_files directory where pickle files are actually saved
        search_directories = ['./analysis_files/', './analysis_output/', './']
        
        for search_dir in search_directories:
            pickle_path = os.path.join(search_dir, pickle_basename)
            if os.path.exists(pickle_path):
                print(f"Found pickle file using y_hat_path: {pickle_path}")
                return pickle_path
        
        print(f"Warning: Expected pickle file not found: {pickle_basename}")
    
    # Fallback to original matching logic
    search_directories = ['./analysis_files/', './analysis_output/', './']
    
    for search_dir in search_directories:
        # Look for pickle files that match scan name specifically
        pattern = f"{search_dir}*{scan_name}*{config_name}*results.pickle"
        pickle_files = glob.glob(pattern)
        
        if not pickle_files:
            # Try with just scan name
            pattern = f"{search_dir}{scan_name}*results.pickle"
            pickle_files = glob.glob(pattern)
        
        if not pickle_files:
            # Try more general patterns
            # Extract key parts of config name for more flexible matching
            config_parts = config_name.split('_')
            if len(config_parts) >= 2:
                key_config = '_'.join(config_parts[:2])  # e.g., 'fiddlercrab_corneas'
                pattern = f"{search_dir}*{scan_name}*{key_config}*results.pickle"
                pickle_files = glob.glob(pattern)
        
        if pickle_files:
            print(f"Found pickle file using fallback method: {pickle_files[0]}")
            return pickle_files[0]
    
    # Debug: show what files are available
    print(f"Debug: Looking for pickle for {config_name} - {scan_name}")
    print("Available pickle files:")
    for search_dir in search_directories:
        available_pickles = glob.glob(f"{search_dir}*results.pickle")
        for pickle_file in available_pickles[:5]:  # Show only first 5 to avoid spam
            print(f"  {pickle_file}")
        if len(available_pickles) > 5:
            print(f"  ... and {len(available_pickles) - 5} more")
    
    print(f"Warning: No pickle file found for {config_name} - {scan_name}")
    return None

def load_data_from_pickle(pickle_path, target_scan_path):
    """Load data from pickle file and find the specific scan."""
    try:
        with open(pickle_path, 'rb') as f:
            data_list = pickle.load(f)
        
        # Find the data entry that matches our target scan
        scan_basename = os.path.basename(target_scan_path)
        for data in data_list:
            if os.path.basename(data['mct_path']) == scan_basename:
                return data
        
        print(f"Warning: Scan {scan_basename} not found in pickle {pickle_path}")
        return None
    except Exception as e:
        print(f"Error loading pickle {pickle_path}: {e}")
        return None

def generate_short_filename(y_hat_path, mct_path, config_path):
    """Generate a shorter, unique filename to avoid filesystem limits."""
    # Extract key components
    scan_name = os.path.splitext(os.path.basename(mct_path))[0].replace('-image', '')
    config_name = os.path.splitext(os.path.basename(config_path))[0]
    
    # Create a short hash of the full y_hat_path for uniqueness
    full_path_hash = hashlib.md5(y_hat_path.encode()).hexdigest()[:8]
    
    # Combine into a shorter filename
    short_name = f"{scan_name}_{config_name}_{full_path_hash}"
    
    # Truncate if still too long (keep under 200 chars to leave room for suffixes)
    if len(short_name) > 150:
        short_name = short_name[:150]
    
    return short_name

def plot_and_edit_points(data, model_info):
    """Loads the mct image and allows interactive point editing in Napari."""
    mct = load_nifti_image(data['mct_path'])
    
    # Generate shorter filename
    short_filename = generate_short_filename(data['y_hat_path'], model_info['mct_path'], model_info['config_path'])
    
    print(f"\nEditing points for: {short_filename}")
    print(f"Config: {model_info['config_path']}")
    print(f"Current metrics - TPs: {model_info['num_tps']}, FPs: {model_info['num_fps']}, FNs: {model_info['num_fns']}")
    print(f"Current F1: {model_info['f1']:.3f}")
    print("\nInstructions:")
    print("- Select points and use buttons to move them between categories")
    print("- Green = True Positives, Red = False Positives, Blue = False Negatives")
    print("- NEW: Press 'z' to smart-convert FPs to TPs + auto-remove nearest FNs (within 50 units)!")
    print("- Press 's' to save when finished")
    print("- Close the viewer window when done")
    
    viewer = napari.Viewer()
    # Add the 3D image to the viewer
    viewer.add_image(mct, name='MCT Image')

    # Add true positives as points in green
    tps_layer = viewer.add_points(data['tps'], size=5, edge_color='green', face_color='green', name='True Positives')

    # Add false positives as points in red
    fps_layer = viewer.add_points(data['fps'], size=5, edge_color='red', face_color='red', name='False Positives')

    # Add false negatives as points in blue
    fns_layer = viewer.add_points(data['fns'], size=5, edge_color='blue', face_color='blue', name='False Negatives')

    # Enable editing for the points
    tps_layer.editable = True
    fps_layer.editable = True
    fns_layer.editable = True

    def move_selected_points(from_layer, to_layer):
            """Move selected points from one layer to another."""
            if len(from_layer.selected_data) == 0:
                print(f"No points selected in {from_layer.name}")
                return
            
            # Get selected point indices and coordinates
            selected_indices = list(from_layer.selected_data)
            selected_points = from_layer.data[selected_indices]
            
            # Add points to destination layer
            if len(to_layer.data) == 0:
                to_layer.data = selected_points
            else:
                to_layer.data = np.vstack([to_layer.data, selected_points])
            
            # Remove points from source layer (in reverse order to maintain indices)
            remaining_data = []
            for i, point in enumerate(from_layer.data):
                if i not in selected_indices:
                    remaining_data.append(point)
            
            if len(remaining_data) > 0:
                from_layer.data = np.array(remaining_data)
            else:
                from_layer.data = np.array([]).reshape(0, 3)
            
            # Clear selection
            from_layer.selected_data = set()
            print(f"Moved {len(selected_indices)} points from {from_layer.name} to {to_layer.name}")

    def find_nearest_points(target_points, candidate_points, max_distance=50.0):
        """Find the nearest candidate point for each target point within max_distance."""
        if len(candidate_points) == 0 or len(target_points) == 0:
            print(f"Debug: No candidate points ({len(candidate_points)}) or target points ({len(target_points)})")
            return []
        
        target_points = np.array(target_points)
        candidate_points = np.array(candidate_points)
        
        print(f"Debug: Looking for nearest FNs to {len(target_points)} FPs among {len(candidate_points)} FNs")
        
        nearest_indices = []
        for i, target_point in enumerate(target_points):
            # Calculate distances to all candidate points
            distances = np.sqrt(np.sum((candidate_points - target_point) ** 2, axis=1))
            
            # Find closest point
            min_distance_idx = np.argmin(distances)
            min_distance = distances[min_distance_idx]
            
            print(f"Debug: FP {i+1} at {target_point}, closest FN at distance {min_distance:.2f}")
            
            if min_distance <= max_distance:
                nearest_indices.append(min_distance_idx)
                print(f"  -> Will remove FN {min_distance_idx} (distance {min_distance:.2f})")
            else:
                nearest_indices.append(None)  # No nearby point found
                print(f"  -> No FN close enough (distance {min_distance:.2f} > {max_distance})")
        
        return nearest_indices

    def smart_fp_to_tp_with_fn_removal(viewer):
        """Convert selected FPs to TPs and automatically remove nearest FNs."""
        if len(fps_layer.selected_data) == 0:
            print("No False Positives selected")
            return
        
        # Get selected FP indices and coordinates
        selected_fp_indices = list(fps_layer.selected_data)
        selected_fp_points = fps_layer.data[selected_fp_indices]
        
        if len(selected_fp_points) == 0:
            return
        
        # Find nearest FNs to the selected FPs
        if len(fns_layer.data) > 0:
            nearest_fn_indices = find_nearest_points(selected_fp_points, fns_layer.data, max_distance=50.0)
            
            # Remove the nearest FNs (process in reverse order to maintain indices)
            fns_to_remove = [idx for idx in nearest_fn_indices if idx is not None]
            fns_to_remove = sorted(set(fns_to_remove), reverse=True)  # Remove duplicates and sort in reverse
            
            # Create new FN data without the removed points
            remaining_fns = []
            for i, point in enumerate(fns_layer.data):
                if i not in fns_to_remove:
                    remaining_fns.append(point)
            
            if len(remaining_fns) > 0:
                fns_layer.data = np.array(remaining_fns)
            else:
                fns_layer.data = np.array([]).reshape(0, 3)
            
            print(f"Removed {len(fns_to_remove)} nearest False Negatives")
        
        # Move FPs to TPs
        if len(tps_layer.data) == 0:
            tps_layer.data = selected_fp_points
        else:
            tps_layer.data = np.vstack([tps_layer.data, selected_fp_points])
        
        # Remove selected FPs from FP layer
        remaining_fps = []
        for i, point in enumerate(fps_layer.data):
            if i not in selected_fp_indices:
                remaining_fps.append(point)
        
        if len(remaining_fps) > 0:
            fps_layer.data = np.array(remaining_fps)
        else:
            fps_layer.data = np.array([]).reshape(0, 3)
        
        # Clear selection
        fps_layer.selected_data = set()
        print(f"Converted {len(selected_fp_indices)} False Positives to True Positives and removed nearest False Negatives")

    # Add buttons for switching between categories
    @viewer.bind_key('q')
    def fp_to_tp(viewer):
        """Move selected False Positives to True Positives (press 'q')"""
        move_selected_points(fps_layer, tps_layer)

    @viewer.bind_key('w')
    def tp_to_fp(viewer):
        """Move selected True Positives to False Positives (press 'w')"""
        move_selected_points(tps_layer, fps_layer)

    @viewer.bind_key('e')
    def fn_to_tp(viewer):
        """Move selected False Negatives to True Positives (press 'e')"""
        move_selected_points(fns_layer, tps_layer)

    @viewer.bind_key('r')
    def tp_to_fn(viewer):
        """Move selected True Positives to False Negatives (press 'r')"""
        move_selected_points(tps_layer, fns_layer)

    @viewer.bind_key('t')
    def fp_to_fn(viewer):
        """Move selected False Positives to False Negatives (press 't')"""
        move_selected_points(fps_layer, fns_layer)

    @viewer.bind_key('y')
    def fn_to_fp(viewer):
        """Move selected False Negatives to False Positives (press 'y')"""
        move_selected_points(fns_layer, fps_layer)

    @viewer.bind_key('z')
    def auto_fp_to_tp(viewer):
        """Smart convert: FP→TP + auto-remove nearest FNs (press 'z')"""
        smart_fp_to_tp_with_fn_removal(viewer)

    @viewer.bind_key('s')
    def save_points(viewer):
        """Save edited points when 's' is pressed."""
        os.makedirs('corrected_results', exist_ok=True)
        
        np.savetxt(f'corrected_results/{short_filename}_cleaned_tps.csv', 
                  tps_layer.data, delimiter=',', header='x,y,z', comments='')
        np.savetxt(f'corrected_results/{short_filename}_cleaned_fps.csv', 
                  fps_layer.data, delimiter=',', header='x,y,z', comments='')
        np.savetxt(f'corrected_results/{short_filename}_cleaned_fns.csv', 
                  fns_layer.data, delimiter=',', header='x,y,z', comments='')
        print(f"Edited points saved to corrected_results/ directory")

        # Print keyboard shortcuts
        print("\nKeyboard Shortcuts:")
        print("  Press 'q': Move selected False Positives → True Positives")
        print("  Press 'w': Move selected True Positives → False Positives") 
        print("  Press 'e': Move selected False Negatives → True Positives")
        print("  Press 'r': Move selected True Positives → False Negatives")
        print("  Press 't': Move selected False Positives → False Negatives")
        print("  Press 'y': Move selected False Negatives → False Positives")
        print("  Press 'z': SMART: FP→TP + auto-remove nearest FNs (within 50 units) (RECOMMENDED!)")
        print("  Press 's': Save corrections")
        print("\nHow to use:")
        print("1. Select a point layer (click on layer name)")
        print("2. Select points (click on them while in select mode)")
        print("3. Press the appropriate letter key to move them")
        print("\nMost efficient: 'z' (auto FP→TP with FN removal within 50 units) or 'e' (FN→TP)")
        print("Standard: 'q' (FP→TP) - use when you don't want auto FN removal")
    
    # Run the napari event loop
    napari.run()

def load_cleaned_points(short_filename):
    """Load the cleaned points from CSV files using short filename."""
    try:
        tps_path = f'corrected_results/{short_filename}_cleaned_tps.csv'
        fps_path = f'corrected_results/{short_filename}_cleaned_fps.csv'
        fns_path = f'corrected_results/{short_filename}_cleaned_fns.csv'
        
        def load_points_safe(path):
            if os.path.exists(path) and os.path.getsize(path) > 5:
                try:
                    points = np.loadtxt(path, delimiter=',', skiprows=1, ndmin=2)
                    if points.size == 0:
                        return np.array([]).reshape(0, 3)
                    return points
                except:
                    return np.array([]).reshape(0, 3)
            return np.array([]).reshape(0, 3)
        
        tps = load_points_safe(tps_path)
        fps = load_points_safe(fps_path)
        fns = load_points_safe(fns_path)
        
        return tps.tolist(), fps.tolist(), fns.tolist()
    except Exception as e:
        print(f"Error loading cleaned points for {short_filename}: {e}")
        return None, None, None

def calculate_point_overlap(points1, points2, tolerance=1.0):
    """Calculate the percentage of points from points1 that exist in points2 within tolerance."""
    if len(points1) == 0:
        return 0.0, 0
    
    if len(points2) == 0:
        return 0.0, 0
    
    points1 = np.array(points1)
    points2 = np.array(points2)
    
    matches = 0
    for p1 in points1:
        # Calculate distances to all points in points2
        distances = np.sqrt(np.sum((points2 - p1) ** 2, axis=1))
        # Check if any point is within tolerance
        if np.min(distances) <= tolerance:
            matches += 1
    
    percentage = (matches / len(points1)) * 100
    return percentage, matches

def calculate_file_similarity(file1_data, file2_data):
    """Calculate similarity between two files based on their characteristics."""
    # Extract key features for comparison
    def extract_features(data):
        features = {}
        # Basic counts
        features['num_tps'] = data.get('num_tps', 0)
        features['num_fps'] = data.get('num_fps', 0)
        features['num_fns'] = data.get('num_fns', 0)
        features['total_points'] = features['num_tps'] + features['num_fps'] + features['num_fns']
        
        # Ratios
        if features['total_points'] > 0:
            features['tp_ratio'] = features['num_tps'] / features['total_points']
            features['fp_ratio'] = features['num_fps'] / features['total_points']
            features['fn_ratio'] = features['num_fns'] / features['total_points']
        else:
            features['tp_ratio'] = 0
            features['fp_ratio'] = 0
            features['fn_ratio'] = 0
        
        # Performance metrics
        features['f1'] = data.get('f1', 0)
        features['precision'] = data.get('precision', 0)
        features['recall'] = data.get('recall', 0)
        
        # Config characteristics
        config_name = os.path.basename(data.get('config_path', ''))
        if 'cornea' in config_name.lower():
            features['structure_type'] = 'cornea'
        elif 'rhabdom' in config_name.lower():
            features['structure_type'] = 'rhabdom'
        else:
            features['structure_type'] = 'other'
        
        # Species/scan type
        scan_name = os.path.basename(data.get('mct_path', '')).lower()
        if 'fiddlercrab' in scan_name or 'dampieri' in scan_name or 'flammula' in scan_name:
            features['species'] = 'fiddlercrab'
        elif 'paraphronima' in scan_name:
            features['species'] = 'paraphronima'
        elif 'cystisoma' in scan_name:
            features['species'] = 'cystisoma'
        else:
            features['species'] = 'other'
        
        return features
    
    features1 = extract_features(file1_data)
    features2 = extract_features(file2_data)
    
    # Calculate similarity scores for different aspects
    similarity_scores = {}
    
    # Count similarity (normalized difference)
    count_diff = abs(features1['total_points'] - features2['total_points'])
    max_count = max(features1['total_points'], features2['total_points'], 1)
    similarity_scores['count'] = 1 - (count_diff / max_count)
    
    # Ratio similarity
    ratio_similarity = 0
    for ratio_type in ['tp_ratio', 'fp_ratio', 'fn_ratio']:
        ratio_similarity += 1 - abs(features1[ratio_type] - features2[ratio_type])
    similarity_scores['ratio'] = ratio_similarity / 3
    
    # Performance similarity
    perf_similarity = 0
    for metric in ['f1', 'precision', 'recall']:
        perf_similarity += 1 - abs(features1[metric] - features2[metric])
    similarity_scores['performance'] = perf_similarity / 3
    
    # Structure type similarity (exact match)
    similarity_scores['structure'] = 1.0 if features1['structure_type'] == features2['structure_type'] else 0.0
    
    # Species similarity (exact match)
    similarity_scores['species'] = 1.0 if features1['species'] == features2['species'] else 0.5
    
    # Weighted average (giving more weight to structure and species match)
    weights = {
        'count': 0.15,
        'ratio': 0.15,
        'performance': 0.20,
        'structure': 0.30,
        'species': 0.20
    }
    
    total_similarity = sum(similarity_scores[key] * weights[key] for key in weights)
    
    return total_similarity, similarity_scores

def find_best_matching_corrections(target_data, target_short_filename, target_points, n_matches=5):
    """Find the best matching corrected files for an uncorrected file, prioritizing TP spatial overlap."""
    if not os.path.exists('corrected_results'):
        return []
    
    # Get all available corrections
    correction_files = glob.glob('corrected_results/*_cleaned_tps.csv')
    available_corrections = []
    
    for file in correction_files:
        short_filename = os.path.basename(file).replace('_cleaned_tps.csv', '')
        # Skip if this is the same file we're trying to correct
        if short_filename == target_short_filename:
            continue
        available_corrections.append(short_filename)
    
    if not available_corrections:
        return []
    
    # Calculate similarities with all available corrections
    matches = []
    
    for correction_filename in available_corrections:
        # Try to load the corrected points to get basic stats
        tps, fps, fns = load_cleaned_points(correction_filename)
        
        if tps is not None:
            # Calculate TP spatial overlap (most important factor)
            tp_overlap_pct, tp_overlap_count = calculate_point_overlap(
                target_points['tps'], tps, tolerance=0.5)
            
            # Create a data structure similar to target_data for comparison
            corrected_data = {
                'num_tps': len(tps),
                'num_fps': len(fps),
                'num_fns': len(fns),
                'config_path': target_data.get('config_path', ''),
                'mct_path': correction_filename  # Use filename as proxy for scan info
            }
            
            # Calculate metrics
            precision = len(tps) / (len(tps) + len(fps)) if (len(tps) + len(fps)) > 0 else 0
            recall = len(tps) / (len(tps) + len(fns)) if (len(tps) + len(fns)) > 0 else 0
            f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
            
            corrected_data['precision'] = precision
            corrected_data['recall'] = recall
            corrected_data['f1'] = f1
            
            # Calculate similarity (secondary factor)
            total_similarity, detail_scores = calculate_file_similarity(target_data, corrected_data)
            
            # Create composite score prioritizing TP spatial overlap (70% weight) over similarity (30% weight)
            composite_score = (tp_overlap_pct / 100.0) * 0.7 + total_similarity * 0.3
            
            matches.append({
                'filename': correction_filename,
                'tp_overlap_pct': tp_overlap_pct,
                'tp_overlap_count': tp_overlap_count,
                'composite_score': composite_score,
                'similarity': total_similarity,
                'detail_scores': detail_scores,
                'stats': {
                    'tps': len(tps),
                    'fps': len(fps),
                    'fns': len(fns),
                    'f1': f1
                },
                'tps_coords': tps,
                'fps_coords': fps,
                'fns_coords': fns
            })
    
    # Sort by composite score (prioritizing TP spatial overlap) and return top n matches
    matches.sort(key=lambda x: x['composite_score'], reverse=True)
    return matches[:n_matches]

def visualize_template_comparison(mct_image, original_data, template_data, template_name):
    """Show visual comparison of original points vs template points side by side."""
    print(f"\n🔍 Visual Preview: Original vs Template ({template_name})")
    print("="*60)
    print("Instructions:")
    print("- LEFT window: Original uncorrected points")
    print("- RIGHT window: Template corrected points")
    print("- Compare the point distributions visually")
    print("- Close BOTH windows when done reviewing")
    print("="*60)
    
    # Create viewer for original points
    viewer_original = napari.Viewer(title="ORIGINAL (Uncorrected)")
    viewer_original.add_image(mct_image, name='MCT Image')
    
    # Add original points
    if len(original_data['tps']) > 0:
        viewer_original.add_points(original_data['tps'], size=5, edge_color='green', 
                                  face_color='green', name=f"TPs ({len(original_data['tps'])})")
    if len(original_data['fps']) > 0:
        viewer_original.add_points(original_data['fps'], size=5, edge_color='red', 
                                  face_color='red', name=f"FPs ({len(original_data['fps'])})")
    if len(original_data['fns']) > 0:
        viewer_original.add_points(original_data['fns'], size=5, edge_color='blue', 
                                  face_color='blue', name=f"FNs ({len(original_data['fns'])})")
    
    # Create viewer for template points
    viewer_template = napari.Viewer(title=f"TEMPLATE: {template_name}")
    viewer_template.add_image(mct_image, name='MCT Image')
    
    # Add template points
    if len(template_data['tps']) > 0:
        viewer_template.add_points(template_data['tps'], size=5, edge_color='green', 
                                  face_color='green', name=f"TPs ({len(template_data['tps'])})")
    if len(template_data['fps']) > 0:
        viewer_template.add_points(template_data['fps'], size=5, edge_color='red', 
                                  face_color='red', name=f"FPs ({len(template_data['fps'])})")
    if len(template_data['fns']) > 0:
        viewer_template.add_points(template_data['fns'], size=5, edge_color='blue', 
                                  face_color='blue', name=f"FNs ({len(template_data['fns'])})")
    
    # Try to position windows side by side (if possible with current Napari version)
    try:
        # Try the older API first
        viewer_original.window.qt_viewer.move(100, 100)
        viewer_template.window.qt_viewer.move(800, 100)
    except (AttributeError, FutureWarning):
        # If that doesn't work, windows will just open in default positions
        pass
    
    # Sync camera state between viewers
    def sync_cameras():
        """Synchronize camera state between the two viewers."""
        # Get camera state from original viewer
        camera_state = {
            'center': viewer_original.camera.center,
            'zoom': viewer_original.camera.zoom,
            'angles': viewer_original.camera.angles,
        }
        # Apply to template viewer
        viewer_template.camera.center = camera_state['center']
        viewer_template.camera.zoom = camera_state['zoom']
        viewer_template.camera.angles = camera_state['angles']
    
    # Connect camera events for synchronization
    @viewer_original.camera.events.center.connect
    def _on_center_change(event):
        viewer_template.camera.center = event.value
    
    @viewer_original.camera.events.zoom.connect
    def _on_zoom_change(event):
        viewer_template.camera.zoom = event.value
    
    @viewer_original.camera.events.angles.connect
    def _on_angles_change(event):
        viewer_template.camera.angles = event.value
    
    # Run the napari event loop
    napari.run()

def select_correction_template(matches, target_info, data, mct_image):
    """Interactive terminal interface to select a correction template with visual preview."""
    print("\n" + "="*80)
    print("CORRECTION TEMPLATE SELECTION")
    print("="*80)
    print(f"\nTarget file: {os.path.basename(target_info.get('mct_path', 'Unknown'))}")
    print(f"Config: {os.path.basename(target_info.get('config_path', 'Unknown'))}")
    print(f"Current stats - TPs: {target_info.get('num_tps', 0)}, "
          f"FPs: {target_info.get('num_fps', 0)}, "
          f"FNs: {target_info.get('num_fns', 0)}, "
          f"F1: {target_info.get('f1', 0):.3f}")
    
    if not matches:
        print("\nNo existing corrections found to use as templates.")
        return None
    
    print("\n" + "-"*80)
    print("Available correction templates (sorted by TP spatial overlap):")
    print("-"*80)
    
    for i, match in enumerate(matches, 1):
        print(f"\n{i}. {match['filename']}")
        print(f"   🎯 TP Spatial Overlap: {match.get('tp_overlap_pct', 0):.1f}% ({match.get('tp_overlap_count', 0)}/{target_info.get('num_tps', 0)}) points match")
        print(f"   📊 Composite Score: {match.get('composite_score', 0)*100:.1f}%")
        print(f"   🔗 Overall Similarity: {match['similarity']*100:.1f}%")
        print(f"   📈 Stats: TPs={match['stats']['tps']}, FPs={match['stats']['fps']}, "
              f"FNs={match['stats']['fns']}, F1={match['stats']['f1']:.3f}")
    
    print(f"\n{len(matches)+1}. Start from scratch (no template)")
    print(f"{len(matches)+2}. Show more templates")
    print("-"*80)
    
    while True:
        try:
            choice = input(f"\nSelect option to preview (1-{len(matches)+2}): ").strip()
            choice_num = int(choice)
            
            if 1 <= choice_num <= len(matches):
                selected = matches[choice_num - 1]
                print(f"\n📋 Template Details: {selected['filename']}")
                
                # Prepare data for overlap calculation
                original_points = {
                    'tps': np.array(data.get('tps', [])),
                    'fps': np.array(data.get('fps', [])),
                    'fns': np.array(data.get('fns', []))
                }
                
                template_points = {
                    'tps': np.array(selected['tps_coords']),
                    'fps': np.array(selected['fps_coords']),
                    'fns': np.array(selected['fns_coords'])
                }
                
                # Calculate exact point overlaps
                print("📊 Calculating point overlaps...")
                tp_overlap_pct, tp_overlap_count = calculate_point_overlap(
                    original_points['tps'], template_points['tps'], tolerance=0.5)
                fp_overlap_pct, fp_overlap_count = calculate_point_overlap(
                    original_points['fps'], template_points['fps'], tolerance=0.5)
                fn_overlap_pct, fn_overlap_count = calculate_point_overlap(
                    original_points['fns'], template_points['fns'], tolerance=0.5)
                
                print("\n🎯 Exact Point Overlap (within 0.5 unit tolerance):")
                print(f"   True Positives:  {tp_overlap_pct:5.1f}% ({tp_overlap_count}/{len(original_points['tps'])}) points match")
                print(f"   False Positives: {fp_overlap_pct:5.1f}% ({fp_overlap_count}/{len(original_points['fps'])}) points match")
                print(f"   False Negatives: {fn_overlap_pct:5.1f}% ({fn_overlap_count}/{len(original_points['fns'])}) points match")
                
                # Calculate overall overlap
                total_original = len(original_points['tps']) + len(original_points['fps']) + len(original_points['fns'])
                total_matches = tp_overlap_count + fp_overlap_count + fn_overlap_count
                overall_overlap = (total_matches / total_original * 100) if total_original > 0 else 0
                print(f"   Overall:         {overall_overlap:5.1f}% ({total_matches}/{total_original}) points match")
                
                # Ask if they want visual comparison
                print(f"\n✓ Template: {selected['filename']}")
                print(f"  Point overlap: {overall_overlap:.1f}% of original points exist in template")
                print(f"  This will use the corrected points from this file as a starting point.")
                
                while True:
                    choice = input("\nWhat would you like to do?\n  (y) Use this template for editing\n  (a) Accept template as-is (skip editing)\n  (v) Visual comparison first\n  (r) Review another template\n  (n) Cancel\nChoice: ").lower().strip()
                    
                    if choice in ['y', 'yes']:
                        return selected
                    elif choice in ['a', 'accept', 'as-is']:
                        # Mark this template for use as-is
                        selected['use_as_is'] = True
                        return selected
                    elif choice in ['v', 'visual', 'view']:
                        print("\nOpening visual comparison...")
                        visualize_template_comparison(mct_image, original_points, template_points, selected['filename'])
                        print(f"\nAfter visual review of {selected['filename']}:")
                        final_choice = input("  Confirm selection? (y/n): ").lower().strip()
                        if final_choice in ['y', 'yes']:
                            return selected
                        else:
                            print("  Selection cancelled.")
                            break  # Go back to template selection
                    elif choice in ['r', 'review']:
                        print("  Let's review another option...")
                        break  # Go back to template selection
                    elif choice in ['n', 'no', 'cancel']:
                        print("  Selection cancelled. Please choose again.")
                        break  # Go back to template selection
                    else:
                        print("  Invalid choice. Please enter 'y', 'a', 'v', 'r', or 'n'.")
            elif choice_num == len(matches) + 1:
                print("\n✓ Starting from scratch (no template)")
                confirm = input("  Confirm starting from scratch? (y/n): ").lower().strip()
                if confirm in ['y', 'yes']:
                    return None
                else:
                    print("  Selection cancelled. Please choose again.")
            elif choice_num == len(matches) + 2:
                print("\n🔍 Showing more templates...")
                # Get more templates (expand to 15 total)
                target_points_for_more = {
                    'tps': np.array(data.get('tps', [])),
                    'fps': np.array(data.get('fps', [])),
                    'fns': np.array(data.get('fns', []))
                }
                more_matches = find_best_matching_corrections(target_info, 
                    os.path.basename(target_info.get('mct_path', '')), target_points_for_more, n_matches=15)
                
                if len(more_matches) > len(matches):
                    return select_correction_template(more_matches, target_info, data, mct_image)
                else:
                    print("No additional templates available.")
                    print("Returning to current selection...")
            else:
                print(f"Invalid choice. Please enter a number between 1 and {len(matches)+2}")
        except ValueError:
            print(f"Invalid input. Please enter a number between 1 and {len(matches)+2}")
        except KeyboardInterrupt:
            print("\n\nSelection cancelled by user.")
            return None

def recalculate_metrics(tps, fps, fns):
    """Recalculate performance metrics from corrected points."""
    num_tps = len(tps) if tps is not None else 0
    num_fps = len(fps) if fps is not None else 0
    num_fns = len(fns) if fns is not None else 0
    
    # Calculate precision, recall, and F1 score
    precision = num_tps / (num_tps + num_fps) if (num_tps + num_fps) > 0 else 0
    recall = num_tps / (num_tps + num_fns) if (num_tps + num_fns) > 0 else 0
    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    
    return {
        'num_tps_corrected': num_tps,
        'num_fps_corrected': num_fps,
        'num_fns_corrected': num_fns,
        'precision_corrected': precision,
        'recall_corrected': recall,
        'f1_corrected': f1_score
    }

def process_model(row, use_template_matching=True):
    """Process a single model from the best models table."""
    print(f"\n{'='*60}")
    print(f"Processing: {os.path.basename(row['mct_path'])}")
    print(f"Config: {os.path.basename(row['config_path'])}")
    print(f"{'='*60}")
    
    # Find the corresponding pickle file
    pickle_path = find_pickle_file_for_model(row)
    if not pickle_path:
        print("Skipping - no pickle file found")
        return None
    
    # Load data from pickle
    data = load_data_from_pickle(pickle_path, row['mct_path'])
    if not data:
        print("Skipping - data not found in pickle")
        return None
    
    # Print current metrics
    print(f"Current Performance:")
    print(f"  True Positives: {row['num_tps']}")
    print(f"  False Positives: {row['num_fps']}")
    print(f"  False Negatives: {row['num_fns']}")
    print(f"  Recall: {row['recall']:.3f}")
    print(f"  F1 Score: {row['f1']:.3f}")
    
    # Generate short filename
    short_filename = generate_short_filename(data['y_hat_path'], row['mct_path'], row['config_path'])
    
    # Check if corrections already exist
    corrected_files_exist = all(os.path.exists(f'corrected_results/{short_filename}_cleaned_{suffix}.csv') 
                               for suffix in ['tps', 'fps', 'fns'])
    
    if not corrected_files_exist:
        # Check if we should use template matching
        template_selected = None
        if use_template_matching:
            # Load MCT image for visual comparison
            mct_image = load_nifti_image(data['mct_path'])
            
            # Prepare target data for similarity matching
            target_data = {
                'num_tps': row['num_tps'],
                'num_fps': row['num_fps'],
                'num_fns': row['num_fns'],
                'precision': row['precision'],
                'recall': row['recall'],
                'f1': row['f1'],
                'config_path': row['config_path'],
                'mct_path': row['mct_path']
            }
            
            # Prepare target points for comparison
            target_points = {
                'tps': np.array(data.get('tps', [])),
                'fps': np.array(data.get('fps', [])),
                'fns': np.array(data.get('fns', []))
            }
            
            # Find best matching corrections
            matches = find_best_matching_corrections(target_data, short_filename, target_points, n_matches=5)
            
            if matches:
                # Let user select a template with visual preview
                template_selected = select_correction_template(matches, target_data, data, mct_image)
                
                if template_selected:
                    print(f"\nUsing template: {template_selected['filename']}")
                    print(f"Template stats - TPs: {template_selected['stats']['tps']}, "
                          f"FPs: {template_selected['stats']['fps']}, "
                          f"FNs: {template_selected['stats']['fns']}")
                    
                    # Use template coordinates as starting point
                    data['tps'] = np.array(template_selected['tps_coords'])
                    data['fps'] = np.array(template_selected['fps_coords'])
                    data['fns'] = np.array(template_selected['fns_coords'])
                    
                    # Check if user wants to use template as-is
                    if template_selected.get('use_as_is', False):
                        print("\n✅ Using template as-is, skipping manual editing...")
                        # Save the template points directly as corrected results
                        os.makedirs('corrected_results', exist_ok=True)
                        np.savetxt(f'corrected_results/{short_filename}_cleaned_tps.csv', 
                                  data['tps'], delimiter=',', header='x,y,z', comments='')
                        np.savetxt(f'corrected_results/{short_filename}_cleaned_fps.csv', 
                                  data['fps'], delimiter=',', header='x,y,z', comments='')
                        np.savetxt(f'corrected_results/{short_filename}_cleaned_fns.csv', 
                                  data['fns'], delimiter=',', header='x,y,z', comments='')
                        print(f"Template corrections saved directly to corrected_results/")
                    else:
                        print("\nOpening Napari with template corrections as starting point...")
                        plot_and_edit_points(data, row)
                        print("Napari session completed.")
                else:
                    print("\nStarting from scratch (no template selected)")
                    print(f"\nOpening Napari for manual correction...")
                    plot_and_edit_points(data, row)
                    print("Napari session completed.")
            else:
                print("\nNo correction templates available. Starting from scratch.")
                print(f"\nOpening Napari for manual correction...")
                plot_and_edit_points(data, row)
                print("Napari session completed.")
    else:
        print(f"Using existing corrections...")
    
    # Load corrected points and recalculate metrics
    tps_corrected, fps_corrected, fns_corrected = load_cleaned_points(short_filename)
    
    if tps_corrected is not None:
        corrected_metrics = recalculate_metrics(tps_corrected, fps_corrected, fns_corrected)
        
        print(f"\nCorrected Performance:")
        print(f"  True Positives: {corrected_metrics['num_tps_corrected']}")
        print(f"  False Positives: {corrected_metrics['num_fps_corrected']}")
        print(f"  False Negatives: {corrected_metrics['num_fns_corrected']}")
        print(f"  Precision: {corrected_metrics['precision_corrected']:.3f}")
        print(f"  Recall: {corrected_metrics['recall_corrected']:.3f}")
        print(f"  F1 Score: {corrected_metrics['f1_corrected']:.3f}")
        
        improvement = corrected_metrics['f1_corrected'] - row['f1']
        print(f"  F1 Improvement: {improvement:+.3f}")
        
        # Combine original and corrected data
        result = {
            'scan_name': os.path.basename(row['mct_path']),
            'config_name': os.path.basename(row['config_path']),
            'scan_path': row['mct_path'],
            'config_path': row['config_path'],
            'short_filename': short_filename,  # Add this for reference
            
            # Original metrics
            'original_tps': row['num_tps'],
            'original_fps': row['num_fps'],
            'original_fns': row['num_fns'],
            'original_precision': row['precision'],
            'original_recall': row['recall'],
            'original_f1': row['f1'],
            
            # Corrected metrics
            'corrected_tps': corrected_metrics['num_tps_corrected'],
            'corrected_fps': corrected_metrics['num_fps_corrected'],
            'corrected_fns': corrected_metrics['num_fns_corrected'],
            'corrected_precision': corrected_metrics['precision_corrected'],
            'corrected_recall': corrected_metrics['recall_corrected'],
            'corrected_f1': corrected_metrics['f1_corrected'],
            
            # Improvements
            'f1_improvement': improvement,
            'precision_improvement': corrected_metrics['precision_corrected'] - row['precision'],
            'recall_improvement': corrected_metrics['recall_corrected'] - row['recall'],
            
            # Ground truth data
            'corrected_tps_coords': tps_corrected,
            'corrected_fps_coords': fps_corrected,
            'corrected_fns_coords': fns_corrected,
        }
        
        return result
    else:
        print("Failed to load corrected data")
        return None

def save_results_to_excel(results, output_path="corrected_models_results.xlsx"):
    """Save all results to a comprehensive Excel file."""
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # Sheet 1: Summary metrics
    ws_summary = wb.active
    ws_summary.title = "Summary_Metrics"
    
    # Prepare summary data
    summary_data = []
    for result in results:
        summary_data.append({
            'Scan': result['scan_name'],
            'Config': result['config_name'],
            'Original_TPs': result['original_tps'],
            'Original_FPs': result['original_fps'],
            'Original_FNs': result['original_fns'],
            'Original_Precision': result['original_precision'],
            'Original_Recall': result['original_recall'],
            'Original_F1': result['original_f1'],
            'Corrected_TPs': result['corrected_tps'],
            'Corrected_FPs': result['corrected_fps'],
            'Corrected_FNs': result['corrected_fns'],
            'Corrected_Precision': result['corrected_precision'],
            'Corrected_Recall': result['corrected_recall'],
            'Corrected_F1': result['corrected_f1'],
            'F1_Improvement': result['f1_improvement'],
            'Precision_Improvement': result['precision_improvement'],
            'Recall_Improvement': result['recall_improvement']
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # Write summary to sheet
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)
    
    # Sheet 2: Corrected True Positives coordinates
    ws_tps = wb.create_sheet("Corrected_True_Positives")
    ws_tps.append(['Scan', 'Config', 'X', 'Y', 'Z'])
    
    for result in results:
        scan_name = result['scan_name']
        config_name = result['config_name']
        for coord in result['corrected_tps_coords']:
            if len(coord) >= 3:
                ws_tps.append([scan_name, config_name, coord[0], coord[1], coord[2]])
    
    # Sheet 3: Corrected False Positives coordinates
    ws_fps = wb.create_sheet("Corrected_False_Positives")
    ws_fps.append(['Scan', 'Config', 'X', 'Y', 'Z'])
    
    for result in results:
        scan_name = result['scan_name']
        config_name = result['config_name']
        for coord in result['corrected_fps_coords']:
            if len(coord) >= 3:
                ws_fps.append([scan_name, config_name, coord[0], coord[1], coord[2]])
    
    # Sheet 4: Corrected False Negatives coordinates
    ws_fns = wb.create_sheet("Corrected_False_Negatives")
    ws_fns.append(['Scan', 'Config', 'X', 'Y', 'Z'])
    
    for result in results:
        scan_name = result['scan_name']
        config_name = result['config_name']
        for coord in result['corrected_fns_coords']:
            if len(coord) >= 3:
                ws_fns.append([scan_name, config_name, coord[0], coord[1], coord[2]])
    
    # Save the workbook
    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")
    
    # Print summary statistics
    print(f"\n{'='*60}")
    print("CORRECTION SUMMARY")
    print(f"{'='*60}")
    print(f"Total models processed: {len(results)}")
    print(f"Average F1 improvement: {summary_df['F1_Improvement'].mean():.3f}")
    print(f"Average precision improvement: {summary_df['Precision_Improvement'].mean():.3f}")
    print(f"Average recall improvement: {summary_df['Recall_Improvement'].mean():.3f}")
    print(f"\nBest F1 improvement: {summary_df['F1_Improvement'].max():.3f}")
    print(f"Worst F1 change: {summary_df['F1_Improvement'].min():.3f}")
    
    return summary_df

def review_corrected_model(row):
    """Review a previously corrected model."""
    print(f"\n{'='*60}")
    print(f"REVIEWING: {os.path.basename(row['mct_path'])}")
    print(f"Config: {os.path.basename(row['config_path'])}")
    print(f"{'='*60}")
    
    # Find the corresponding pickle file
    pickle_path = find_pickle_file_for_model(row)
    if not pickle_path:
        print("Skipping - no pickle file found")
        return None
    
    # Load data from pickle
    data = load_data_from_pickle(pickle_path, row['mct_path'])
    if not data:
        print("Skipping - data not found in pickle")
        return None
    
    # Generate short filename
    short_filename = generate_short_filename(data['y_hat_path'], row['mct_path'], row['config_path'])
    
    # Check if corrections exist
    corrected_files_exist = all(os.path.exists(f'corrected_results/{short_filename}_cleaned_{suffix}.csv') 
                               for suffix in ['tps', 'fps', 'fns'])
    
    if not corrected_files_exist:
        print("No corrections found for this model. Run without --review first.")
        return None
    
    # Load corrected points
    tps_corrected, fps_corrected, fns_corrected = load_cleaned_points(short_filename)
    
    if tps_corrected is None:
        print("Failed to load corrected data")
        return None
    
    # Calculate corrected metrics
    corrected_metrics = recalculate_metrics(tps_corrected, fps_corrected, fns_corrected)
    
    # Display comparison
    print(f"\nORIGINAL vs CORRECTED COMPARISON:")
    print(f"{'Metric':<15} {'Original':<10} {'Corrected':<10} {'Change':<10}")
    print(f"{'-'*45}")
    print(f"{'True Positives':<15} {row['num_tps']:<10} {corrected_metrics['num_tps_corrected']:<10} {corrected_metrics['num_tps_corrected'] - row['num_tps']:+d}")
    print(f"{'False Positives':<15} {row['num_fps']:<10} {corrected_metrics['num_fps_corrected']:<10} {corrected_metrics['num_fps_corrected'] - row['num_fps']:+d}")
    print(f"{'False Negatives':<15} {row['num_fns']:<10} {corrected_metrics['num_fns_corrected']:<10} {corrected_metrics['num_fns_corrected'] - row['num_fns']:+d}")
    print(f"{'Precision':<15} {row['precision']:<10.3f} {corrected_metrics['precision_corrected']:<10.3f} {corrected_metrics['precision_corrected'] - row['precision']:+.3f}")
    print(f"{'Recall':<15} {row['recall']:<10.3f} {corrected_metrics['recall_corrected']:<10.3f} {corrected_metrics['recall_corrected'] - row['recall']:+.3f}")
    print(f"{'F1 Score':<15} {row['f1']:<10.3f} {corrected_metrics['f1_corrected']:<10.3f} {corrected_metrics['f1_corrected'] - row['f1']:+.3f}")
    
    # Ask user if they want to view/edit in Napari
    user_input = input(f"\nWould you like to view/edit this correction in Napari? (y/n): ").lower().strip()
    
    if user_input in ['y', 'yes']:
        review_in_napari(data, row, tps_corrected, fps_corrected, fns_corrected, short_filename)
    
    return {
        'scan_name': os.path.basename(row['mct_path']),
        'config_name': os.path.basename(row['config_path']),
        'short_filename': short_filename,
        'original_metrics': {
            'tps': row['num_tps'],
            'fps': row['num_fps'], 
            'fns': row['num_fns'],
            'precision': row['precision'],
            'recall': row['recall'],
            'f1': row['f1']
        },
        'corrected_metrics': corrected_metrics,
        'improvements': {
            'f1': corrected_metrics['f1_corrected'] - row['f1'],
            'precision': corrected_metrics['precision_corrected'] - row['precision'],
            'recall': corrected_metrics['recall_corrected'] - row['recall']
        }
    }

def review_in_napari(data, model_info, tps_corrected, fps_corrected, fns_corrected, short_filename):
    """Review corrected points in Napari with ability to make additional edits."""
    mct = load_nifti_image(data['mct_path'])
    
    print(f"\nReviewing corrections for: {short_filename}")
    print(f"Current corrected metrics - TPs: {len(tps_corrected)}, FPs: {len(fps_corrected)}, FNs: {len(fns_corrected)}")
    print("\nReview Mode Instructions:")
    print("- Green = Corrected True Positives, Red = Corrected False Positives, Blue = Corrected False Negatives")
    print("- You can still edit points using the same keyboard shortcuts")
    print("- Press 's' to save any additional changes")
    print("- Close the viewer window when done")
    
    viewer = napari.Viewer()
    viewer.add_image(mct, name='MCT Image')

    # Add corrected points
    tps_layer = viewer.add_points(
        np.array(tps_corrected) if len(tps_corrected) > 0 else np.array([]).reshape(0, 3), 
        size=5, edge_color='green', face_color='green', name='Corrected True Positives'
    )
    fps_layer = viewer.add_points(
        np.array(fps_corrected) if len(fps_corrected) > 0 else np.array([]).reshape(0, 3), 
        size=5, edge_color='red', face_color='red', name='Corrected False Positives'
    )
    fns_layer = viewer.add_points(
        np.array(fns_corrected) if len(fns_corrected) > 0 else np.array([]).reshape(0, 3), 
        size=5, edge_color='blue', face_color='blue', name='Corrected False Negatives'
    )

    # Enable editing
    tps_layer.editable = True
    fps_layer.editable = True
    fns_layer.editable = True

    def move_selected_points(from_layer, to_layer):
        """Move selected points from one layer to another."""
        if len(from_layer.selected_data) == 0:
            print(f"No points selected in {from_layer.name}")
            return
        
        selected_indices = list(from_layer.selected_data)
        selected_points = from_layer.data[selected_indices]
        
        if len(to_layer.data) == 0:
            to_layer.data = selected_points
        else:
            to_layer.data = np.vstack([to_layer.data, selected_points])
        
        remaining_data = []
        for i, point in enumerate(from_layer.data):
            if i not in selected_indices:
                remaining_data.append(point)
        
        if len(remaining_data) > 0:
            from_layer.data = np.array(remaining_data)
        else:
            from_layer.data = np.array([]).reshape(0, 3)
        
        from_layer.selected_data = set()
        print(f"Moved {len(selected_indices)} points from {from_layer.name} to {to_layer.name}")

    def find_nearest_points_review(target_points, candidate_points, max_distance=50.0):
        """Find the nearest candidate point for each target point within max_distance."""
        if len(candidate_points) == 0 or len(target_points) == 0:
            print(f"Debug: No candidate points ({len(candidate_points)}) or target points ({len(target_points)})")
            return []
        
        target_points = np.array(target_points)
        candidate_points = np.array(candidate_points)
        
        print(f"Debug: Looking for nearest FNs to {len(target_points)} FPs among {len(candidate_points)} FNs")
        
        nearest_indices = []
        for i, target_point in enumerate(target_points):
            # Calculate distances to all candidate points
            distances = np.sqrt(np.sum((candidate_points - target_point) ** 2, axis=1))
            
            # Find closest point
            min_distance_idx = np.argmin(distances)
            min_distance = distances[min_distance_idx]
            
            print(f"Debug: FP {i+1} at {target_point}, closest FN at distance {min_distance:.2f}")
            
            if min_distance <= max_distance:
                nearest_indices.append(min_distance_idx)
                print(f"  -> Will remove FN {min_distance_idx} (distance {min_distance:.2f})")
            else:
                nearest_indices.append(None)  # No nearby point found
                print(f"  -> No FN close enough (distance {min_distance:.2f} > {max_distance})")
        
        return nearest_indices

    def smart_fp_to_tp_with_fn_removal_review(viewer):
        """Convert selected FPs to TPs and automatically remove nearest FNs."""
        if len(fps_layer.selected_data) == 0:
            print("No False Positives selected")
            return
        
        # Get selected FP indices and coordinates
        selected_fp_indices = list(fps_layer.selected_data)
        selected_fp_points = fps_layer.data[selected_fp_indices]
        
        if len(selected_fp_points) == 0:
            return
        
        # Find nearest FNs to the selected FPs
        if len(fns_layer.data) > 0:
            nearest_fn_indices = find_nearest_points_review(selected_fp_points, fns_layer.data, max_distance=70.0)
            
            # Remove the nearest FNs (process in reverse order to maintain indices)
            fns_to_remove = [idx for idx in nearest_fn_indices if idx is not None]
            fns_to_remove = sorted(set(fns_to_remove), reverse=True)  # Remove duplicates and sort in reverse
            
            # Create new FN data without the removed points
            remaining_fns = []
            for i, point in enumerate(fns_layer.data):
                if i not in fns_to_remove:
                    remaining_fns.append(point)
            
            if len(remaining_fns) > 0:
                fns_layer.data = np.array(remaining_fns)
            else:
                fns_layer.data = np.array([]).reshape(0, 3)
            
            print(f"Removed {len(fns_to_remove)} nearest False Negatives")
        
        # Move FPs to TPs
        if len(tps_layer.data) == 0:
            tps_layer.data = selected_fp_points
        else:
            tps_layer.data = np.vstack([tps_layer.data, selected_fp_points])
        
        # Remove selected FPs from FP layer
        remaining_fps = []
        for i, point in enumerate(fps_layer.data):
            if i not in selected_fp_indices:
                remaining_fps.append(point)
        
        if len(remaining_fps) > 0:
            fps_layer.data = np.array(remaining_fps)
        else:
            fps_layer.data = np.array([]).reshape(0, 3)
        
        # Clear selection
        fps_layer.selected_data = set()
        print(f"Converted {len(selected_fp_indices)} False Positives to True Positives and removed nearest False Negatives")

    # Add the same keyboard shortcuts as in the main editing mode
    @viewer.bind_key('q')
    def fp_to_tp(viewer):
        move_selected_points(fps_layer, tps_layer)

    @viewer.bind_key('w')
    def tp_to_fp(viewer):
        move_selected_points(tps_layer, fps_layer)

    @viewer.bind_key('e')
    def fn_to_tp(viewer):
        move_selected_points(fns_layer, tps_layer)

    @viewer.bind_key('r')
    def tp_to_fn(viewer):
        move_selected_points(tps_layer, fns_layer)

    @viewer.bind_key('t')
    def fp_to_fn(viewer):
        move_selected_points(fps_layer, fns_layer)

    @viewer.bind_key('y')
    def fn_to_fp(viewer):
        move_selected_points(fns_layer, fps_layer)

    @viewer.bind_key('z')
    def auto_fp_to_tp_review(viewer):
        """Smart convert: FP→TP + auto-remove nearest FNs (press 'z')"""
        smart_fp_to_tp_with_fn_removal_review(viewer)

    @viewer.bind_key('s')
    def save_points(viewer):
        """Save edited points when 's' is pressed."""
        os.makedirs('corrected_results', exist_ok=True)
        
        np.savetxt(f'corrected_results/{short_filename}_cleaned_tps.csv', 
                  tps_layer.data, delimiter=',', header='x,y,z', comments='')
        np.savetxt(f'corrected_results/{short_filename}_cleaned_fps.csv', 
                  fps_layer.data, delimiter=',', header='x,y,z', comments='')
        np.savetxt(f'corrected_results/{short_filename}_cleaned_fns.csv', 
                  fns_layer.data, delimiter=',', header='x,y,z', comments='')
        print(f"Updated corrections saved to corrected_results/ directory")

    print("\nKeyboard Shortcuts (same as editing mode):")
    print("  'q': FP→TP, 'w': TP→FP, 'e': FN→TP, 'r': TP→FN, 't': FP→FN, 'y': FN→FP")
    print("  'z': SMART FP→TP + auto-remove nearest FNs (within 50 units)")
    print("  's': Save changes")
    
    # Run the napari event loop
    napari.run()

def list_available_corrections():
    """List all available corrected results."""
    if not os.path.exists('corrected_results'):
        print("No corrected_results directory found.")
        return []
    
    # Find all correction files
    correction_files = glob.glob('corrected_results/*_cleaned_tps.csv')
    available_corrections = []
    
    for file in correction_files:
        short_filename = os.path.basename(file).replace('_cleaned_tps.csv', '')
        available_corrections.append(short_filename)
    
    if available_corrections:
        print(f"\nFound {len(available_corrections)} corrected models:")
        for i, correction in enumerate(available_corrections, 1):
            print(f"  {i}. {correction}")
    else:
        print("No corrected results found.")
    
    return available_corrections

def run_review_mode(best_models_df, filter_config=None):
    """Run the review mode for previously corrected models."""
    print(f"\n{'='*60}")
    print("REVIEW MODE - Examining Previously Corrected Models")
    print(f"{'='*60}")
    
    # List available corrections
    available_corrections = list_available_corrections()
    
    if not available_corrections:
        print("No corrections available to review. Run the main workflow first.")
        return
    
    # Filter models based on config if specified
    if filter_config:
        mask = best_models_df['config_path'].str.contains(filter_config, na=False)
        filtered_df = best_models_df[mask]
    else:
        # Default to fiddlercrab corneas
        mask = best_models_df['config_path'].str.contains('fiddlercrab_corneas', na=False)
        filtered_df = best_models_df[mask]
    
    if len(filtered_df) == 0:
        print("No matching models found for review.")
        return
    
    # Find which models have corrections
    models_with_corrections = []
    for idx, row in filtered_df.iterrows():
        # Find pickle file to get short filename
        pickle_path = find_pickle_file_for_model(row)
        if pickle_path:
            data = load_data_from_pickle(pickle_path, row['mct_path'])
            if data:
                short_filename = generate_short_filename(data['y_hat_path'], row['mct_path'], row['config_path'])
                if short_filename in available_corrections:
                    models_with_corrections.append((row, short_filename))
    
    if not models_with_corrections:
        print("No models with corrections found matching the filter criteria.")
        return
    
    print(f"\nFound {len(models_with_corrections)} models with corrections to review:")
    for i, (row, short_filename) in enumerate(models_with_corrections, 1):
        print(f"  {i}. {os.path.basename(row['mct_path'])} - {os.path.basename(row['config_path'])}")
    
    # Process each model with corrections
    review_results = []
    for row, short_filename in models_with_corrections:
        result = review_corrected_model(row)
        if result:
            review_results.append(result)
    
    # Print summary
    if review_results:
        print(f"\n{'='*60}")
        print("REVIEW SUMMARY")
        print(f"{'='*60}")
        
        total_f1_improvement = sum(r['improvements']['f1'] for r in review_results)
        avg_f1_improvement = total_f1_improvement / len(review_results)
        
        print(f"Models reviewed: {len(review_results)}")
        print(f"Average F1 improvement: {avg_f1_improvement:.3f}")
        print(f"Total F1 improvement: {total_f1_improvement:.3f}")
        
        print(f"\nDetailed Results:")
        print(f"{'Model':<30} {'Original F1':<12} {'Corrected F1':<12} {'Improvement':<12}")
        print(f"{'-'*66}")
        for result in review_results:
            original_f1 = result['original_metrics']['f1']
            corrected_f1 = result['corrected_metrics']['f1_corrected']
            improvement = result['improvements']['f1']
            model_name = f"{result['scan_name'][:15]}...{result['config_name'][:10]}"
            print(f"{model_name:<30} {original_f1:<12.3f} {corrected_f1:<12.3f} {improvement:<+12.3f}")

def main():
    parser = argparse.ArgumentParser(description='Comprehensive model correction workflow')
    parser.add_argument('--csv_path', type=str, default='best_models.csv',
                       help='Path to the CSV file containing best models')
    parser.add_argument('--output_path', type=str, default='corrected_models_results.xlsx',
                       help='Path for the output Excel file')
    parser.add_argument('--filter_config', type=str, default=None,
                       help='Only process models from this config (optional)')
    parser.add_argument('--review', action='store_true',
                       help='Review mode: examine previously corrected models instead of creating new corrections')
    parser.add_argument('--no-template', action='store_true',
                       help='Disable template matching - always start corrections from scratch')
    
    args = parser.parse_args()
    
    # Load best models from CSV
    best_models_df = load_best_models_from_csv(args.csv_path)
    if best_models_df is None:
        print("Failed to load best models. Exiting.")
        return
    
    # Run in review mode if requested
    if args.review:
        run_review_mode(best_models_df, args.filter_config)
        return
    
    # Original workflow continues below...
    # Filter for specified config
    if args.filter_config:
        mask = best_models_df['config_path'].str.contains(args.filter_config, na=False)
        filtered_df = best_models_df[mask]
    else:
        filtered_df = best_models_df
    
    if len(filtered_df) == 0:
        print("No matching models found for correction.")
        if args.filter_config:
            print(f"Filter: {args.filter_config}")
        else:
            print("Filter: fiddlercrab_corneas")
        print("Available configs:")
        for config in best_models_df['config_path'].unique():
            print(f"  {config}")
        return
    
    print(f"Found {len(filtered_df)} models to process")
    
    # Process each model
    results = []
    use_template = not args.no_template
    
    if use_template:
        print("\n📋 Template matching is ENABLED")
        print("   (Use --no-template to disable)")
    else:
        print("\n⚠️  Template matching is DISABLED")
        print("   All corrections will start from scratch")
    
    for idx, row in filtered_df.iterrows():
        print(f"\nProcessing model {idx + 1}/{len(filtered_df)}: {os.path.basename(row['mct_path'])}")
        result = process_model(row, use_template_matching=use_template)
        if result:
            results.append(result)
    
    if results:
        # Save results to Excel
        summary_df = save_results_to_excel(results, args.output_path)
        
        print(f"\nDetailed results:")
        print(summary_df[['Scan', 'Config', 'Original_F1', 'Corrected_F1', 'F1_Improvement']].to_string(index=False))
    else:
        print("No results to save.")

if __name__ == "__main__":
    main() 
